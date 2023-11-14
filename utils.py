import os
from pathlib import Path
from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup
import re
import openpyxl
import requests
import json

class Listing:
    def __init__(self, title="", price="", location="", link="", desc="", condition="", genre="", age_group="", image_path="", image_url="", scrapped=""):
        self.title = title
        self.price = price
        self.location = location
        self.link = link
        self.desc = desc
        self.condition = condition
        self.genre = genre
        self.age_group = age_group
        self.image_path = image_path
        self.image_url = image_url
        self.scrapped = scrapped

class Sheet:
    def __init__(self, file_path, option):
        self.file_path = file_path
        if option == Sheet.option_create:
            self.create_sheet()
        else:
            self.load_sheet()

    def create_sheet(self):
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.sheet.append(["Title", "Price", "Location", "Link", "Description", "Condition", "Genre", "Age Group", "Image Path", "Image URL", "Scrapped"])

    def load_sheet(self):
        self.workbook = openpyxl.load_workbook(self.file_path)
        self.sheet = self.workbook.active

    def add_listing(self, listing):
        self.sheet.append([listing.title, listing.price, listing.location, listing.link, listing.desc, listing.condition, listing.genre, listing.age_group, listing.image_path, listing.image_url, listing.scrapped])

    def contains(self, link):
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            if row[3] == link:  # Assuming the link is in the fourth column (0-based index)
                return True
        return False

    def update_listing(self, new_row):
        for index, row in enumerate(self.sheet.iter_rows(min_row=2, values_only=True), start=2):
            if row[3] == new_row[3]:
                for col_index in range(len(new_row)):
                    self.sheet.cell(row=index, column=col_index + 1, value=new_row[col_index])

    def listing(self):
        scraped_count = 0
        for _ in self.sheet.iter_rows(min_row=2, values_only=True):
            scraped_count += 1
        return scraped_count - 1
    
    def listings(self):
        return self.sheet
    
    def save(self):
        self.workbook.save(self.file_path)

# Function to scroll down a webpage using Playwright
def scroll_page(page, num_of_times=1):
    for _ in range(num_of_times):
        page.keyboard.press("Space")
        page.keyboard.up("Space")

# Function to clean price text (remove extra characters)
def clean_price_text(price):
    return price.strip("$").replace(",", "")

# Function to get the total scrape count from user input
def get_listing_scrape_count():
    try:
        total_scrape_limit = int(input("Enter the total scrape limit: "))
        return total_scrape_limit
    except ValueError:
        print("Invalid input. Please enter a valid number.")
        return get_listing_scrape_count()
    
# Function to get the total scrape count from user input
def get_input_file_name():
    while True:
        filename = input("Enter the input file name : ")

        # Check if the file exists
        if not os.path.exists(filename):
            print("File does not exist. Please provide a valid file name.")
            continue

        # Split the filename into its base name and extension
        base_name, extension = os.path.splitext(filename)
        
        # Check if the extension is ".xlsx" (case insensitive)
        if extension.lower() == ".xlsx":
            return filename
        else:
            print("Invalid file extension. Please provide a valid excel file name.")


# Function to get an available file name for Excel sheets
def find_available_file_name(base_name="output", extension=".xlsx"):
    base_name = base_name + extension
    count = 1
    while os.path.exists(base_name):
        base_name = f"{base_name.split('.')[0]}_{count}{extension}"
        count += 1
    return base_name

# Function to sanitize a string for use as a filename
def sanitize_filename(filename):
    # Remove any characters that aren't alphanumeric or whitespace
    filename = re.sub(r'[^\w\s]', '', filename)
    # Replace spaces with underscores
    filename = filename.replace(' ', '_')
    return filename

# Function to download an image and save it to a specified path
def download_image(image_path, image_url):
    try:
        response = requests.get(image_url, stream=True)
        if response.status_code == 200:
            with open(image_path, 'wb') as file:
                for chunk in response.iter_content(1024):
                    file.write(chunk)
    except Exception as e:
        print(f"Error downloading image: {e}")

# Function to read an Excel sheet and return its contents as a list of dictionaries
def read_excel_sheet(file_path, sheet_name):
    try:
        workbook = load_workbook(file_path)
        sheet = workbook[sheet_name]

        # Get header row
        header = [cell.value for cell in sheet[1]]

        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            item = dict(zip(header, row))
            data.append(item)

        return data

    except Exception as error:
        print(f"Error reading Excel sheet: {error}")
        return []

# Function to write data to an Excel sheet
def write_to_excel_sheet(file_path, sheet_name, data):
    try:
        workbook = load_workbook(file_path)
        sheet = workbook[sheet_name]

        # Write header row
        header = list(data[0].keys())
        sheet.append(header)

        # Write data
        for item in data:
            sheet.append(list(item.values()))

        workbook.save(file_path)
        print(f"Data written to Excel sheet: {sheet_name}")

    except Exception as error:
        print(f"Error writing to Excel sheet: {error}")

# Function to analyze text using the ChatGPT API (implement your ChatGPT API integration here)
def analyze_text_with_chatgpt(text, api_key):
    # Implement your ChatGPT API integration here
    # You can make an API request to analyze the text and return the resultsq
    url = "https://api.openai.com/v1/chat/completions"

    payload = json.dumps({
    "model": "gpt-3.5-turbo",
    "messages": [
        {
        "role": "user",
        "content": text
        }
    ],
    "temperature": 0.7
    })
    headers = {
    'Content-Type': 'application/json',
    'Authorization': f'Bearer {api_key}'
    }

    response = requests.post(url, headers=headers, data=payload)

    return response.text

Sheet.option_create = "create"
Sheet.option_update = "update"